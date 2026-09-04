"""Unit tests for Excel Exporter."""
import os
import openpyxl
import pandas as pd
from src.export.excel_exporter import ExcelReportExporter

def test_excel_export_styling(tmp_path):
    output_file = str(tmp_path / "test_report.xlsx")
    
    results_df = pd.DataFrame({
        'Business_Unit': ['BU1', 'BU2'],
        'Ref2_Invoice_No': ['1001', '1002'],
        'Total_CD_LC': [1500.0, 2000.0],
        'Total_Sales_Value': [1500.0, 1900.0],
        'Amount_Variance': [0.0, 100.0],
        'Posting_Date': ['2026-08-01', '2026-08-02'],
        'Sales_DocDate': ['2026-08-01', '2026-08-02'],
        'Customer_Id': ['5001', '5002'],
        'Mapped_SAP_Code': ['5001', '5002'],
        'SAP_Offset_Account': ['5001', '5002'],
        'Overall_Status': ['Matched', 'Not Matched'],
        'Reconciliation_Remarks': ['MATCHED', 'MISMATCHED: Amount Variance (100.0)'],
        'Format_Used': ['Format 2', 'Format 2']
    })

    exporter = ExcelReportExporter()
    exporter.export(output_file, results_df)

    assert os.path.exists(output_file)

    wb = openpyxl.load_workbook(output_file)
    assert 'Executive Summary' in wb.sheetnames
    assert 'Recon Detailed Results' in wb.sheetnames

    ws_details = wb['Recon Detailed Results']
    # Check Header background (1F4E78)
    assert ws_details.cell(row=1, column=1).fill.start_color.rgb == '001F4E78'
    
    # Check Row 2 (Matched) background (D6FFD6)
    assert ws_details.cell(row=2, column=1).fill.start_color.rgb == '00D6FFD6'

    # Check Row 3 (Not Matched) background (FFD6D6)
    assert ws_details.cell(row=3, column=1).fill.start_color.rgb == '00FFD6D6'


def test_collection_details_are_written_to_collection_sheet(tmp_path):
    output_file = str(tmp_path / 'collection_report.xlsx')
    results_df = pd.DataFrame({
        'Recon_Type': ['Collection', 'Collection'],
        'Bank_Name': ['ICICI', 'ICICI'],
        'Bank_UTR': ['UTR-1', 'UTR-2'],
        'Bank_Amount': [100.0, 200.0],
        'SAP_Amount': [100.0, 150.0],
        'Amount_Variance': [0.0, 50.0],
        'Overall_Status': ['Matched', 'Not Matched'],
        'Reconciliation_Remarks': ['MATCHED', 'MISMATCHED: Amount variance'],
    })

    ExcelReportExporter().export(output_file, results_df)

    workbook = openpyxl.load_workbook(output_file, read_only=True)
    assert workbook['Recon Detailed Results'].max_row == 3
    assert workbook['Collection'].max_row == 3
    assert 'Sales - SAP Data' not in workbook.sheetnames
    assert 'Sales - DB Data' not in workbook.sheetnames


def test_sales_sap_and_db_sheets_and_summary_table(tmp_path):
    output_file = str(tmp_path / 'sales_report.xlsx')
    results_df = pd.DataFrame({
        'Recon_Type': ['Sales', 'Sales', 'Sales', 'Sales'],
        'Business_Unit': ['24', '14', '24', '14'],
        'Ref2_Invoice_No': ['INV01', 'INV02', 'INV03', 'INV04'],
        'Posting_Date': ['2026-08-01', '2026-08-02', '2026-08-03', '2026-08-04'],
        'Sales_DocDate': ['2026-08-01', '2026-08-02', '2026-08-03', '2026-08-04'],
        'Total_CD_LC': [1000.0, 500.0, -200.0, -100.0],
        'Total_Sales_Value': [1000.0, 500.0, -200.0, -100.0],
        'Amount_Variance': [0.0, 0.0, 0.0, 0.0],
        'Customer_Id': ['C1', 'C2', 'C3', 'C4'],
        'Mapped_SAP_Code': ['S1', 'S2', 'S3', 'S4'],
        'SAP_Offset_Account': ['A1', 'A2', 'A3', 'A4'],
        'COGSCostingCode': ['24', '14', '24', '14'],
        'Overall_Status': ['Matched', 'Matched', 'Matched', 'Matched'],
        'Reconciliation_Remarks': ['MATCHED', 'MATCHED', 'MATCHED', 'MATCHED'],
        'Format_Used': ['Format 2', 'Format 2', 'Format 2', 'Format 2'],
    })

    # Exact raw input dataframes as uploaded
    raw_sap_df = pd.DataFrame({
        'Raw_SAP_Col1': [1, 2, 3],
        'Raw_SAP_Col2': ['A', 'B', 'C'],
        'Extra_Audit_Col': ['X', 'Y', 'Z'],
    })
    raw_db_df = pd.DataFrame({
        'Raw_DB_Col1': [10, 20, 30],
        'Raw_DB_Col2': ['D1', 'D2', 'D3'],
        'Audit_Field': [99.9, 88.8, 77.7],
    })
    results_df.attrs['raw_sales_sap'] = raw_sap_df
    results_df.attrs['raw_sales_db'] = raw_db_df

    exporter = ExcelReportExporter()
    exporter.export(output_file, results_df)

    wb = openpyxl.load_workbook(output_file)
    assert 'Executive Summary' in wb.sheetnames
    assert 'Recon Detailed Results' in wb.sheetnames
    assert 'Sales' in wb.sheetnames
    assert 'Sales - SAP Data' in wb.sheetnames
    assert 'Sales - DB Data' in wb.sheetnames

    # Check Sales - SAP Data contains exact raw uploaded columns
    ws_sap = wb['Sales - SAP Data']
    sap_headers = [cell.value for cell in ws_sap[1]]
    assert sap_headers == ['Raw_SAP_Col1', 'Raw_SAP_Col2', 'Extra_Audit_Col']

    # Check Sales - DB Data contains exact raw uploaded columns
    ws_db = wb['Sales - DB Data']
    db_headers = [cell.value for cell in ws_db[1]]
    assert db_headers == ['Raw_DB_Col1', 'Raw_DB_Col2', 'Audit_Field']

    # Check Executive Summary table content
    ws_summary = wb['Executive Summary']
    cell_values = [[cell.value for cell in row] for row in ws_summary.iter_rows()]
    found_title = any('Sales Reconciliation Summary' in str(row) for row in cell_values)
    assert found_title

    # Read summary table using helper
    summary_df = ExcelReportExporter._build_sales_summary_table(results_df)
    assert len(summary_df) == 4
    # Sales BU 14 -> 1 row (INV02), sum=500
    row_sales_14 = summary_df[(summary_df['Particulars'] == 'Sales') & (summary_df['BU'] == '14')].iloc[0]
    assert row_sales_14['Total line item as per DB'] == 1
    assert row_sales_14['DB Amount'] == 500.0

    # Sales BU 24 -> 1 row (INV01), sum=1000
    row_sales_24 = summary_df[(summary_df['Particulars'] == 'Sales') & (summary_df['BU'] == '24')].iloc[0]
    assert row_sales_24['Total line item as per DB'] == 1
    assert row_sales_24['DB Amount'] == 1000.0
    assert row_sales_24['SAP Amount'] == 1000.0


def test_dynamic_bu_sales_summary_table():
    """Verify arbitrary BUs (e.g. 10, 14, 24, BU_Kolkata) are dynamically identified."""
    results_df = pd.DataFrame({
        'Business_Unit': ['10', '14', '24', 'BU_Kolkata', '10'],
        'Total_CD_LC': [100.0, 200.0, 300.0, -50.0, 100.0],
        'Total_Sales_Value': [100.0, 200.0, 300.0, -50.0, 100.0],
        'Amount_Variance': [0.0, 0.0, 0.0, 0.0, 0.0],
    })
    summary_df = ExcelReportExporter._build_sales_summary_table(results_df)
    unique_bus_in_summary = set(summary_df['BU'])
    assert unique_bus_in_summary == {'10', '14', '24', 'BU_Kolkata'}
    # 4 distinct BUs * 2 (Sales and CN) = 8 rows
    assert len(summary_df) == 8

    # Check Sales row for BU 10: 2 line items, sum=200
    row_sales_10 = summary_df[(summary_df['Particulars'] == 'Sales') & (summary_df['BU'] == '10')].iloc[0]
    assert row_sales_10['Total line item as per DB'] == 2
    assert row_sales_10['DB Amount'] == 200.0


def test_collection_sap_and_bank_sheets_and_summary_table(tmp_path):
    """Verify Collection - SAP Data, Collection - Bank Data, and Collection Summary Table in Executive Summary."""
    output_file = str(tmp_path / 'collection_report_styled.xlsx')
    results_df = pd.DataFrame({
        'Recon_Type': ['Collection', 'Collection', 'Collection'],
        'Bank_Name': ['ICICI', 'ICICI', 'SCB'],
        'Bank_Account_Number': ['ICICI - 107505004178', 'ICICI - 107505004792', 'SCB - 45505426790'],
        'Bank_Date': ['2026-08-01', '2026-08-02', '2026-08-03'],
        'Bank_UTR': ['UTR1', 'UTR2', 'UTR3'],
        'SAP_Offset_Account': ['ACC1', 'ACC2', 'ACC3'],
        'Bank_Amount': [1000.0, 2000.0, 1500.0],
        'SAP_Posting_Date': ['2026-08-01', '2026-08-02', '2026-08-03'],
        'SAP_Amount': [1000.0, 1900.0, 1500.0],
        'Amount_Variance': [0.0, 100.0, 0.0],
        'Overall_Status': ['Matched', 'Not Matched', 'Matched'],
        'Reconciliation_Remarks': ['MATCHED', 'MISMATCHED: Amount variance', 'MATCHED'],
    })

    raw_sap_coll = pd.DataFrame({
        'SAP_GL': ['1010202089', '1010202094'],
        'Details': ['UTR1', 'UTR2'],
        'Credit (LC)': [1000.0, 1900.0]
    })
    raw_bank_coll = pd.DataFrame({
        'Tran_ID': ['UTR1', 'UTR2', 'UTR3'],
        'Deposit_Amt': [1000.0, 2000.0, 1500.0]
    })
    results_df.attrs['raw_collection_sap'] = raw_sap_coll
    results_df.attrs['raw_collection_bank'] = raw_bank_coll

    exporter = ExcelReportExporter()
    exporter.export(output_file, results_df)

    wb = openpyxl.load_workbook(output_file)
    assert 'Executive Summary' in wb.sheetnames
    assert 'Recon Detailed Results' in wb.sheetnames
    assert 'Collection' in wb.sheetnames
    assert 'Collection - SAP Data' in wb.sheetnames
    assert 'Collection - Bank Data' in wb.sheetnames
    assert 'Sales - SAP Data' not in wb.sheetnames
    assert 'Sales - DB Data' not in wb.sheetnames

    # Check Collection Summary table builder
    summary_df = ExcelReportExporter._build_collection_summary_table(results_df)
    assert len(summary_df) == 3
    assert list(summary_df.columns) == [
        'Particulars', 'Account Number', 'Total line item as per Bank',
        'Bank Amount', 'SAP Amount', 'Amount Variance'
    ]
    # Check ICICI 107505004178
    row_icici1 = summary_df[summary_df['Account Number'] == '107505004178'].iloc[0]
    assert row_icici1['Particulars'] == 'ICICI'
    assert row_icici1['Total line item as per Bank'] == 1
    assert row_icici1['Bank Amount'] == 1000.0
    assert row_icici1['SAP Amount'] == 1000.0
    assert row_icici1['Amount Variance'] == 0.0

    # Check Executive Summary text contains Collection title
    ws_summary = wb['Executive Summary']
    cell_values = [[cell.value for cell in row] for row in ws_summary.iter_rows()]
    assert any('Collection Reconciliation Summary' in str(row) for row in cell_values)


def test_sales_export_with_nan_bu(tmp_path):
    """Ensure rows with NaN or float Business_Unit values do not cause 'float' object has no attribute 'startswith'."""
    output_file = str(tmp_path / 'nan_bu_report.xlsx')
    import numpy as np
    results_df = pd.DataFrame({
        'Recon_Type': ['Sales', 'Sales', 'Sales'],
        'Business_Unit': [float('nan'), np.nan, None],
        'Ref2_Invoice_No': ['INV01', 'INV02', 'INV03'],
        'Posting_Date': ['2026-08-01', 'Missing in SAP', '2026-08-03'],
        'Sales_DocDate': ['2026-08-01', '2026-08-02', '2026-08-03'],
        'Total_CD_LC': [1000.0, 0.0, 500.0],
        'Total_Sales_Value': [1000.0, 1084746.05, 500.0],
        'Amount_Variance': [0.0, -1084746.05, 0.0],
        'Customer_Id': ['C1', 'C2', 'C3'],
        'Mapped_SAP_Code': ['S1', 'S2', 'S3'],
        'SAP_Offset_Account': ['A1', 'Missing in SAP', 'A3'],
        'Overall_Status': ['Matched', 'Not Matched', 'Matched'],
        'Reconciliation_Remarks': ['MATCHED', 'MISMATCHED: Missing in SAP', 'MATCHED'],
        'Format_Used': ['Format 2', 'Format 2', 'Format 2'],
    })

    exporter = ExcelReportExporter()
    exporter.export(output_file, results_df)

    assert os.path.exists(output_file)
    wb = openpyxl.load_workbook(output_file)
    assert 'Executive Summary' in wb.sheetnames
    assert 'Recon Detailed Results' in wb.sheetnames


