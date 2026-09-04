"""Verification of exported reconciliation workbooks against fresh results."""
from dataclasses import dataclass, field
from numbers import Number
from typing import Iterable, List, Optional

import openpyxl
import pandas as pd

from ..services.recon_engine import process_file_list


@dataclass
class VerificationResult:
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        return not self.errors


def _normalise_value(value):
    if pd.isna(value):
        return ""
    if isinstance(value, Number):
        numeric = round(float(value), 6)
        return str(int(numeric)) if numeric.is_integer() else str(numeric)
    return str(value).strip()


def _normalise_frame(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.map(_normalise_value)


def _expected_kpis(results: pd.DataFrame) -> dict:
    matched = int((results["Overall_Status"] == "Matched").sum())
    sap_column = "Total_CD_LC" if "Total_CD_LC" in results else "SAP_Amount"
    other_column = "Total_Sales_Value" if "Total_Sales_Value" in results else "Bank_Amount"
    return {
        "Total Records Reconciled": len(results),
        "Fully Matched Count": matched,
        "Mismatched Count": len(results) - matched,
        "Match Rate (%)": round(matched / len(results) * 100, 1) if len(results) else 0,
        "Total SAP Amount": round(results[sap_column].sum(), 2),
        "Total Sales/Bank Amount": round(results[other_column].sum(), 2),
        "Total Net Variance": round(results["Amount_Variance"].sum(), 2),
    }


def _check_result_invariants(frame: pd.DataFrame, name: str, result: VerificationResult) -> None:
    required = {"Overall_Status", "Reconciliation_Remarks", "Amount_Variance"}
    missing = required - set(frame.columns)
    if missing:
        result.errors.append(f"{name}: missing required columns: {sorted(missing)}")
        return
    invalid_status = set(frame["Overall_Status"].dropna().unique()) - {"Matched", "Not Matched"}
    if invalid_status:
        result.errors.append(f"{name}: invalid Overall_Status values: {sorted(invalid_status)}")
    status_from_remarks = frame["Reconciliation_Remarks"].fillna("").astype(str).str.startswith("MATCHED")
    if not status_from_remarks.equals(frame["Overall_Status"].eq("Matched")):
        result.errors.append(f"{name}: Overall_Status disagrees with Reconciliation_Remarks")
    expected = pd.Series(float("nan"), index=frame.index)
    if {"Total_CD_LC", "Total_Sales_Value"}.issubset(frame.columns):
        sales_rows = frame["Total_CD_LC"].notna() & frame["Total_Sales_Value"].notna()
        expected.loc[sales_rows] = frame.loc[sales_rows, "Total_CD_LC"] - frame.loc[sales_rows, "Total_Sales_Value"]
    if {"Bank_Amount", "SAP_Amount"}.issubset(frame.columns):
        collection_rows = frame["Bank_Amount"].notna() & frame["SAP_Amount"].notna()
        expected.loc[collection_rows] = frame.loc[collection_rows, "Bank_Amount"] - frame.loc[collection_rows, "SAP_Amount"]
    actual = pd.to_numeric(frame["Amount_Variance"], errors="coerce")
    checked = expected.notna()
    if checked.any() and not (actual[checked].round(2).to_numpy() == expected[checked].round(2).to_numpy()).all():
        result.errors.append(f"{name}: Amount_Variance does not match its input amounts")


def verify_workbook(
    report_path: str,
    input_files: Optional[Iterable[str]] = None,
    mode: str = "Auto Detect",
    recon_model: str = "Auto",
) -> VerificationResult:
    """Verify report structure, formulas, and optionally exact fresh output."""
    result = VerificationResult()
    try:
        workbook = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    except Exception as exc:
        result.errors.append(f"Could not open report: {exc}")
        return result
    try:
        detailed_sheet_name = (
            "Recon Detailed Results" if "Recon Detailed Results" in workbook.sheetnames
            else "Detailed Results"
        )
        missing_sheets = {"Executive Summary"} - set(workbook.sheetnames)
        if detailed_sheet_name not in workbook.sheetnames:
            missing_sheets.add("Recon Detailed Results")
        if missing_sheets:
            result.errors.append(f"Missing required sheets: {sorted(missing_sheets)}")
            return result
        try:
            details = pd.read_excel(report_path, sheet_name=detailed_sheet_name)
            summary = pd.read_excel(report_path, sheet_name="Executive Summary", nrows=1)
        except Exception as exc:
            result.errors.append(f"Could not read report tables: {exc}")
            return result
        _check_result_invariants(details, detailed_sheet_name, result)
        if "Recon_Type" in details.columns:
            for recon_type, frame in details.groupby("Recon_Type", dropna=False):
                sheet_name = str(recon_type)[:31]
                if sheet_name not in workbook.sheetnames:
                    result.errors.append(f"Missing per-type sheet: {sheet_name}")
                elif not _normalise_frame(pd.read_excel(report_path, sheet_name=sheet_name)).equals(
                    _normalise_frame(frame.reset_index(drop=True))
                ):
                    result.errors.append(f"Per-type sheet differs from {detailed_sheet_name}: {sheet_name}")
        if len(summary) != 1:
            result.errors.append(f"Executive Summary must contain exactly one data row, found {len(summary)}")
        elif not details.empty:
            for column, expected in _expected_kpis(details).items():
                actual = summary.iloc[0].get(column)
                if _normalise_value(actual) != _normalise_value(expected):
                    result.errors.append(f"Executive Summary {column} is {actual!r}, expected {expected!r}")
        if input_files:
            try:
                fresh = process_file_list(list(input_files), mode=mode, recon_model=recon_model)
            except Exception as exc:
                result.errors.append(f"Fresh reconciliation failed: {exc}")
            else:
                if list(details.columns) != list(fresh.columns):
                    result.errors.append(f"{detailed_sheet_name} columns differ from fresh reconciliation output")
                elif not _normalise_frame(details).equals(_normalise_frame(fresh)):
                    result.errors.append(f"{detailed_sheet_name} data differs from fresh reconciliation output")
        else:
            result.warnings.append("No input files supplied; freshness was not verified")
        return result
    finally:
        workbook.close()