"""Excel Report Exporter with openpyxl styling — separate Sales and Collection sheets."""
from typing import Callable, List, Optional

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..core.constants import STYLE_CONFIG

def _sanitize_cell_value(value):
    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')):
        return f"'{value}"
    return value


class ExcelReportExporter:
    """Exports structured reconciliation results and KPI summaries to styled Excel workbooks."""

    # Columns that belong to the SAP side of a Sales reconciliation
    _SAP_SIDE_COLS = [
        'Business_Unit', 'RefId_Ref1', 'Ref2_Invoice_No', 'Reference',
        'Posting_Date', 'Total_CD_LC', 'SAP_Offset_Account',
        'Mapped_SAP_Code', 'Format_Used',
    ]
    # Columns that belong to the DB side of a Sales reconciliation
    _DB_SIDE_COLS = [
        'Business_Unit', 'RefId_Ref1', 'Ref2_Invoice_No', 'Reference',
        'Sales_DocDate', 'Total_Sales_Value', 'Customer_Id', 'Retailer_Customer_Id',
        'COGSCostingCode', 'Format_Used',
    ]

    def __init__(self, style_config: Optional[dict] = None):
        self.config = style_config or STYLE_CONFIG

    # ------------------------------------------------------------------
    # Sales executive summary table builder
    # ------------------------------------------------------------------
    @staticmethod
    def _build_sales_summary_table(sales_df: pd.DataFrame) -> pd.DataFrame:
        """Build a Sales/CN × BU breakdown table for the Executive Summary.

        Rows are classified as **Sales** (positive amount) or **CN** (negative amount).
        BU is dynamically determined from all distinct Business_Unit values in the data.
        """
        if sales_df.empty:
            return pd.DataFrame()

        # Determine the amount column available
        sap_col = 'Total_CD_LC' if 'Total_CD_LC' in sales_df.columns else None
        db_col = 'Total_Sales_Value' if 'Total_Sales_Value' in sales_df.columns else None
        var_col = 'Amount_Variance' if 'Amount_Variance' in sales_df.columns else None

        sales_df = sales_df.copy()

        # Classify Sales vs CN by sign of amount (negative -> CN, positive/zero -> Sales)
        if sap_col and db_col:
            sap_amt = pd.to_numeric(sales_df[sap_col], errors='coerce').fillna(0)
            db_amt = pd.to_numeric(sales_df[db_col], errors='coerce').fillna(0)
            is_cn = (sap_amt < 0) | (db_amt < 0)
            sales_df['_Particulars'] = np.where(is_cn, 'CN', 'Sales')
        elif sap_col:
            sap_amt = pd.to_numeric(sales_df[sap_col], errors='coerce').fillna(0)
            sales_df['_Particulars'] = np.where(sap_amt < 0, 'CN', 'Sales')
        elif db_col:
            db_amt = pd.to_numeric(sales_df[db_col], errors='coerce').fillna(0)
            sales_df['_Particulars'] = np.where(db_amt < 0, 'CN', 'Sales')
        else:
            sales_df['_Particulars'] = 'Sales'

        # Extract all distinct BU values dynamically from SAP records
        bu_col = 'Business_Unit'
        if bu_col in sales_df.columns:
            cleaned_bus = sales_df[bu_col].fillna('').astype(str).str.strip()
            sales_df['_BU_Clean'] = cleaned_bus
            excluded_bus = {'nan', 'none', 'null', '<na>', '', 'missing in sap', 'missing', 'n/a'}
            valid_bus = [b for b in cleaned_bus.unique() if b and b.lower() not in excluded_bus]
            if not valid_bus:
                valid_bus = ['N/A']
            else:
                def sort_key(x):
                    try:
                        return (0, int(x))
                    except ValueError:
                        return (1, str(x))
                valid_bus = sorted(valid_bus, key=sort_key)
        else:
            sales_df['_BU_Clean'] = 'N/A'
            valid_bus = ['N/A']

        rows: List[dict] = []
        for particulars in ('Sales', 'CN'):
            for bu in valid_bus:
                subset = sales_df[
                    (sales_df['_Particulars'] == particulars)
                    & (sales_df['_BU_Clean'] == bu)
                ]
                rows.append({
                    'Particulars': particulars,
                    'BU': bu,
                    'Total line item as per DB': len(subset),
                    'DB Amount': round(float(pd.to_numeric(subset[db_col], errors='coerce').fillna(0).sum()), 2) if db_col else 0.0,
                    'SAP Amount': round(float(pd.to_numeric(subset[sap_col], errors='coerce').fillna(0).sum()), 2) if sap_col else 0.0,
                    'Amount Variance': round(float(pd.to_numeric(subset[var_col], errors='coerce').fillna(0).sum()), 2) if var_col else 0.0,
                })

        return pd.DataFrame(rows)

    # ------------------------------------------------------------------
    # Collection executive summary table builder
    # ------------------------------------------------------------------
    @staticmethod
    def _build_collection_summary_table(coll_df: pd.DataFrame) -> pd.DataFrame:
        """Build a Collection breakdown table by Bank Name and Account Number.

        Columns: Particulars (Bank Name), Account Number, Total line item as per Bank,
        Bank Amount, SAP Amount, Amount Variance.
        """
        if coll_df.empty:
            return pd.DataFrame()

        import re

        coll_df = coll_df.copy()
        bank_name_col = 'Bank_Name' if 'Bank_Name' in coll_df.columns else None
        bank_acc_col = 'Bank_Account_Number' if 'Bank_Account_Number' in coll_df.columns else None
        db_col = 'Bank_Amount' if 'Bank_Amount' in coll_df.columns else None
        sap_col = 'SAP_Amount' if 'SAP_Amount' in coll_df.columns else None
        var_col = 'Amount_Variance' if 'Amount_Variance' in coll_df.columns else None

        coll_df['_Bank'] = coll_df[bank_name_col].fillna('Bank').astype(str).str.strip() if bank_name_col else 'Bank'

        def clean_acc(val):
            s = str(val).strip() if val is not None and not pd.isna(val) else ''
            if not s or s.lower() in ('nan', 'none', 'null', '<na>'):
                return 'N/A'
            m = re.search(r'\d{6,}', s)
            if m:
                return m.group(0)
            return s

        coll_df['_Account'] = coll_df[bank_acc_col].apply(clean_acc) if bank_acc_col else 'N/A'

        groups = coll_df.groupby(['_Bank', '_Account'], sort=False)
        rows: List[dict] = []
        for (bank, acc), subset in groups:
            db_amt = round(float(pd.to_numeric(subset[db_col], errors='coerce').fillna(0).sum()), 2) if db_col else 0.0
            sap_amt = round(float(pd.to_numeric(subset[sap_col], errors='coerce').fillna(0).sum()), 2) if sap_col else 0.0
            var_amt = round(float(pd.to_numeric(subset[var_col], errors='coerce').fillna(0).sum()), 2) if var_col else round(db_amt - sap_amt, 2)
            rows.append({
                'Particulars': bank,
                'Account Number': acc,
                'Total line item as per Bank': len(subset),
                'Bank Amount': db_amt,
                'SAP Amount': sap_amt,
                'Amount Variance': var_amt,
            })

        return pd.DataFrame(rows)

    # ------------------------------------------------------------------
    # Split Sales data into SAP-side and DB-side DataFrames
    # ------------------------------------------------------------------
    @staticmethod
    def _split_sales_sides(sales_df: pd.DataFrame):
        """Return (sap_df, db_df) containing only the relevant columns."""
        sap_cols = [c for c in ExcelReportExporter._SAP_SIDE_COLS if c in sales_df.columns]
        db_cols = [c for c in ExcelReportExporter._DB_SIDE_COLS if c in sales_df.columns]
        return sales_df[sap_cols].copy(), sales_df[db_cols].copy()

    def export(
        self,
        save_path: str,
        results_df: pd.DataFrame,
        kpi_summary: Optional[pd.DataFrame] = None,
        progress_callback: Optional[Callable[[str, int, int], None]] = None,
        raw_sales_sap: Optional[pd.DataFrame] = None,
        raw_sales_db: Optional[pd.DataFrame] = None,
        raw_collection_sap: Optional[pd.DataFrame] = None,
        raw_collection_bank: Optional[pd.DataFrame] = None,
    ) -> None:
        def report(stage: str, current: int = 0, total: int = 0) -> None:
            if progress_callback:
                progress_callback(stage, current, total)

        report("Preparing Excel report")

        if kpi_summary is None:
            sap_total = (
                results_df['Total_CD_LC'].sum()
                if 'Total_CD_LC' in results_df
                else results_df.get('SAP_Amount', pd.Series(dtype=float)).sum()
            )
            sales_total = (
                results_df['Total_Sales_Value'].sum()
                if 'Total_Sales_Value' in results_df
                else results_df.get('Bank_Amount', pd.Series(dtype=float)).sum()
            )
            var_total = results_df['Amount_Variance'].sum() if 'Amount_Variance' in results_df else 0.0
            total_count = len(results_df)
            matched_count = int((results_df['Overall_Status'] == 'Matched').sum())
            kpi_summary = pd.DataFrame([{
                'Total Records Reconciled': total_count,
                'Fully Matched Count': matched_count,
                'Mismatched Count': total_count - matched_count,
                'Match Rate (%)': round(matched_count / total_count * 100, 1) if total_count else 0,
                'Total SAP Amount': round(sap_total, 2),
                'Total Sales/Bank Amount': round(sales_total, 2),
                'Total Net Variance': round(var_total, 2),
            }])

        # Style helpers
        red_fill = PatternFill(start_color=self.config['red_fill'], end_color=self.config['red_fill'], fill_type='solid')
        green_fill = PatternFill(start_color=self.config['green_fill'], end_color=self.config['green_fill'], fill_type='solid')
        header_fill = PatternFill(start_color=self.config['header_fill'], end_color=self.config['header_fill'], fill_type='solid')
        header_font = Font(color=self.config['header_font_color'], bold=True)
        thin_side = Side(style='thin', color=self.config['border_color'])
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        def _style_sheet(ws: object, df_len: int, sheet_label: str) -> None:
            """Apply header style, frozen pane, autofilter, row colours, and column widths."""
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions

            # Header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Find Overall_Status column index (0-indexed within row tuple)
            status_col_idx = None
            for idx, cell in enumerate(ws[1]):
                if str(cell.value or '').strip() == 'Overall_Status':
                    status_col_idx = idx
                    break

            total_rows = max(ws.max_row - 1, 1)
            row_num = 0
            for row in ws.iter_rows(min_row=2):
                row_num += 1
                status_val = str(row[status_col_idx].value if status_col_idx is not None else '').strip()
                is_matched = status_val.lower() == 'matched'
                row_fill = green_fill if is_matched else red_fill
                row_font_color = self.config['green_font_color'] if is_matched else self.config['red_font_color']
                for idx, cell in enumerate(row):
                    cell.border = thin_border
                    cell.fill = row_fill
                    if idx == status_col_idx:
                        cell.font = Font(color=row_font_color, bold=True)
                if row_num % max(total_rows // 10, 1) == 0:
                    report(f"Styling {sheet_label}", row_num, total_rows)

            # Auto column widths
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 3, 12), 60)

            report(f"Styling {sheet_label}", total_rows, total_rows)

        def _style_plain_sheet(ws: object, sheet_label: str) -> None:
            """Style a data sheet that has no Overall_Status column (e.g. exact SAP/DB data)."""
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border

            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 3, 12), 60)

            report(f"Styling {sheet_label}", 1, 1)

        # ----------------------------------------------------------
        # Extract Sales & Collection data for summary tables & raw sheets
        # ----------------------------------------------------------
        sales_df = pd.DataFrame()
        coll_df = pd.DataFrame()
        if 'Recon_Type' in results_df.columns:
            sales_df = results_df[results_df['Recon_Type'] == 'Sales'].copy()
            coll_df = results_df[results_df['Recon_Type'] == 'Collection'].copy()
        elif 'Total_CD_LC' in results_df.columns:
            sales_df = results_df.copy()
        elif 'Bank_UTR' in results_df.columns:
            coll_df = results_df.copy()

        sales_summary_table = self._build_sales_summary_table(sales_df) if not sales_df.empty else pd.DataFrame()
        coll_summary_table = self._build_collection_summary_table(coll_df) if not coll_df.empty else pd.DataFrame()

        # Raw Sales Data (exact uploaded data)
        raw_sap_s = raw_sales_sap if raw_sales_sap is not None else results_df.attrs.get('raw_sales_sap')
        raw_db_s = raw_sales_db if raw_sales_db is not None else results_df.attrs.get('raw_sales_db')

        if raw_sap_s is not None and isinstance(raw_sap_s, pd.DataFrame) and not raw_sap_s.empty:
            sales_sap_side_df = raw_sap_s.copy()
        elif not sales_df.empty:
            sales_sap_side_df, _ = self._split_sales_sides(sales_df)
        else:
            sales_sap_side_df = pd.DataFrame()

        if raw_db_s is not None and isinstance(raw_db_s, pd.DataFrame) and not raw_db_s.empty:
            sales_db_side_df = raw_db_s.copy()
        elif not sales_df.empty:
            _, sales_db_side_df = self._split_sales_sides(sales_df)
        else:
            sales_db_side_df = pd.DataFrame()

        # Raw Collection Data (exact uploaded data)
        raw_sap_c = raw_collection_sap if raw_collection_sap is not None else results_df.attrs.get('raw_collection_sap')
        raw_bank_c = raw_collection_bank if raw_collection_bank is not None else results_df.attrs.get('raw_collection_bank')

        if raw_sap_c is not None and isinstance(raw_sap_c, pd.DataFrame) and not raw_sap_c.empty:
            coll_sap_side_df = raw_sap_c.copy()
        else:
            coll_sap_side_df = pd.DataFrame()

        if raw_bank_c is not None and isinstance(raw_bank_c, pd.DataFrame) and not raw_bank_c.empty:
            coll_bank_side_df = raw_bank_c.copy()
        else:
            coll_bank_side_df = pd.DataFrame()

        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            report("Writing Executive Summary sheet", 0, 1)
            kpi_summary.to_excel(writer, sheet_name='Executive Summary', index=False)

            report("Writing Recon Detailed Results sheet", 0, 1)
            results_df.to_excel(writer, sheet_name='Recon Detailed Results', index=False)

            # Per-type sheets (Sales / Collection)
            if 'Recon_Type' in results_df.columns:
                for recon_type, frame in results_df.groupby('Recon_Type', sort=False):
                    sheet_name = str(recon_type)[:31] or 'Results'
                    report(f"Writing {sheet_name} sheet", 0, 1)
                    frame.to_excel(writer, sheet_name=sheet_name, index=False)

            # Sales - SAP Data and Sales - DB Data sheets
            if not sales_sap_side_df.empty:
                report("Writing Sales - SAP Data sheet", 0, 1)
                sales_sap_side_df.to_excel(writer, sheet_name='Sales - SAP Data', index=False)
            if not sales_db_side_df.empty:
                report("Writing Sales - DB Data sheet", 0, 1)
                sales_db_side_df.to_excel(writer, sheet_name='Sales - DB Data', index=False)

            # Collection - SAP Data and Collection - Bank Data sheets
            if not coll_sap_side_df.empty:
                report("Writing Collection - SAP Data sheet", 0, 1)
                coll_sap_side_df.to_excel(writer, sheet_name='Collection - SAP Data', index=False)
            if not coll_bank_side_df.empty:
                report("Writing Collection - Bank Data sheet", 0, 1)
                coll_bank_side_df.to_excel(writer, sheet_name='Collection - Bank Data', index=False)

            wb = writer.book

            # ----------------------------------------------------------
            # Style Executive Summary — KPI row + Summary tables
            # ----------------------------------------------------------
            ws_summary = wb['Executive Summary']
            for col_idx in range(1, ws_summary.max_column + 1):
                cell = ws_summary.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in ws_summary.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws_summary.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 14)

            def _write_section_table(title: str, table_df: pd.DataFrame) -> None:
                if table_df.empty:
                    return
                start_row = ws_summary.max_row + 2
                title_cell = ws_summary.cell(row=start_row, column=1, value=title)
                title_cell.font = Font(bold=True, size=12)
                start_row += 1

                summary_headers = list(table_df.columns)
                for col_idx, header in enumerate(summary_headers, 1):
                    cell = ws_summary.cell(row=start_row, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                for row_idx, (_, row_data) in enumerate(table_df.iterrows(), start_row + 1):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws_summary.cell(row=row_idx, column=col_idx, value=_sanitize_cell_value(value))
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                for col_idx, header in enumerate(summary_headers, 1):
                    col_letter = get_column_letter(col_idx)
                    current_width = ws_summary.column_dimensions[col_letter].width or 0
                    new_width = max(len(header) + 3, 14)
                    ws_summary.column_dimensions[col_letter].width = max(current_width, new_width)

            # Write the Sales summary table below the KPI row
            if not sales_summary_table.empty:
                _write_section_table('Sales Reconciliation Summary', sales_summary_table)

            # Write the Collection summary table below
            if not coll_summary_table.empty:
                _write_section_table('Collection Reconciliation Summary', coll_summary_table)

            # Style all data sheets
            sheets_to_style = ['Recon Detailed Results']
            if 'Recon_Type' in results_df.columns:
                for rt in results_df['Recon_Type'].dropna().unique():
                    sn = str(rt)[:31]
                    if sn in wb.sheetnames:
                        sheets_to_style.append(sn)

            for sheet_name in sheets_to_style:
                if sheet_name in wb.sheetnames:
                    df_rows = (
                        len(results_df) if sheet_name == 'Recon Detailed Results'
                        else len(results_df[results_df['Recon_Type'] == sheet_name])
                    )
                    _style_sheet(wb[sheet_name], df_rows, sheet_name)

            # Style the plain raw data sheets (no status colouring)
            for plain_sheet in ('Sales - SAP Data', 'Sales - DB Data', 'Collection - SAP Data', 'Collection - Bank Data'):
                if plain_sheet in wb.sheetnames:
                    _style_plain_sheet(wb[plain_sheet], plain_sheet)

        report("Excel export complete", 1, 1)


def save_styled_reconciliation_excel(
    save_path: str, kpi_summary: pd.DataFrame, results_df: pd.DataFrame
) -> None:
    """Helper wrapper for backward compatibility."""
    ExcelReportExporter().export(save_path, results_df, kpi_summary)
